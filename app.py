import base64
import io
import json
import os
import re

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from pypdf import PdfReader

from backend.chatbot import (build_agent_summary, build_system_prompt,
                             find_comparable_quotes, parse_line_items, parse_quote)
from backend.llm_client import DEPLOYMENT, get_client
from backend.quote_analyzer import analyze_quote, get_dataframe, load_quotes

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AquaCapital — Quote Intelligence",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Global CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ---------- font & base ---------- */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ---------- hide default streamlit chrome ---------- */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 0 !important; }

/* ---------- hero banner ---------- */
.hero {
    background: linear-gradient(135deg, #003B6F 0%, #005F9E 50%, #0097A7 100%);
    border-radius: 16px;
    padding: 32px 40px 28px 40px;
    margin-bottom: 24px;
    position: relative;
    overflow: hidden;
}
.hero::before {
    content: "";
    position: absolute;
    top: -60px; right: -60px;
    width: 260px; height: 260px;
    border-radius: 50%;
    background: rgba(255,255,255,0.06);
}
.hero::after {
    content: "";
    position: absolute;
    bottom: -80px; right: 120px;
    width: 320px; height: 320px;
    border-radius: 50%;
    background: rgba(0,151,167,0.18);
}
.hero-title {
    color: #ffffff;
    font-size: 28px;
    font-weight: 700;
    letter-spacing: -0.5px;
    margin: 0 0 4px 0;
}
.hero-sub {
    color: #B3D9F0;
    font-size: 14px;
    font-weight: 400;
    margin: 0 0 20px 0;
}
.hero-pills { display: flex; gap: 10px; flex-wrap: wrap; }
.hero-pill {
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.25);
    color: #ffffff;
    font-size: 12px;
    font-weight: 500;
    padding: 4px 14px;
    border-radius: 20px;
    backdrop-filter: blur(4px);
}

/* ---------- KPI cards ---------- */
.kpi-grid { display: flex; gap: 14px; margin-bottom: 20px; flex-wrap: wrap; }
.kpi-card {
    flex: 1; min-width: 140px;
    border-radius: 12px;
    padding: 18px 20px 14px 20px;
    position: relative;
    overflow: hidden;
}
.kpi-card.blue  { background: linear-gradient(135deg, #003B6F, #0072BC); }
.kpi-card.teal  { background: linear-gradient(135deg, #006064, #0097A7); }
.kpi-card.green { background: linear-gradient(135deg, #1B5E20, #2E7D32); }
.kpi-card.amber { background: linear-gradient(135deg, #E65100, #F57C00); }
.kpi-card::after {
    content: "";
    position: absolute;
    bottom: -20px; right: -20px;
    width: 90px; height: 90px;
    border-radius: 50%;
    background: rgba(255,255,255,0.08);
}
.kpi-icon { font-size: 22px; margin-bottom: 8px; }
.kpi-label {
    color: rgba(255,255,255,0.75);
    font-size: 11px;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-bottom: 4px;
}
.kpi-value {
    color: #ffffff;
    font-size: 22px;
    font-weight: 700;
    line-height: 1;
}
.kpi-delta {
    color: rgba(255,255,255,0.6);
    font-size: 11px;
    margin-top: 4px;
}

/* ---------- section headers ---------- */
.section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 20px 0 12px 0;
}
.section-icon {
    width: 32px; height: 32px;
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 16px;
    flex-shrink: 0;
}
.section-icon.blue  { background: #E3F2FD; }
.section-icon.teal  { background: #E0F7FA; }
.section-icon.green { background: #E8F5E9; }
.section-icon.amber { background: #FFF3E0; }
.section-title {
    font-size: 15px;
    font-weight: 600;
    color: #1A237E;
    margin: 0;
}
.section-divider {
    height: 2px;
    background: linear-gradient(90deg, #0072BC, #0097A7, transparent);
    border-radius: 2px;
    margin-bottom: 16px;
}

/* ---------- metric cards (analysis) ---------- */
.metric-row { display: flex; gap: 12px; margin-bottom: 18px; }
.metric-box {
    flex: 1;
    border-radius: 12px;
    padding: 16px 18px;
    text-align: center;
    border: 1px solid;
}
.metric-box.neutral { background: #F0F7FF; border-color: #B3D4F0; }
.metric-box.good    { background: #F0FFF4; border-color: #86EFAC; }
.metric-box.warn    { background: #FFFBEB; border-color: #FCD34D; }
.metric-box.bad     { background: #FFF0F0; border-color: #FCA5A5; }
.metric-box-label { font-size: 11px; color: #64748B; font-weight: 500; text-transform: uppercase; letter-spacing: 0.6px; }
.metric-box-value { font-size: 22px; font-weight: 700; color: #1E293B; margin: 4px 0 2px; }
.metric-box-delta { font-size: 12px; font-weight: 600; }
.delta-pos { color: #16A34A; }
.delta-neg { color: #DC2626; }
.delta-neu { color: #64748B; }

/* ---------- agent summary card ---------- */
.agent-summary-card {
    background: linear-gradient(135deg, #EFF6FF 0%, #E0F2FE 100%);
    border: 1px solid #BAE6FD;
    border-left: 5px solid #0072BC;
    border-radius: 14px;
    padding: 20px 22px;
    margin-bottom: 20px;
}
.agent-summary-header {
    display: flex; align-items: center; gap: 10px; margin-bottom: 12px;
}
.agent-summary-title { font-size: 14px; font-weight: 700; color: #1A237E; }
.agent-summary-badge {
    background: linear-gradient(135deg, #0072BC, #0097A7);
    color: white; font-size: 10px; font-weight: 700;
    padding: 2px 9px; border-radius: 12px; letter-spacing: 0.5px;
}
.agent-summary-body { color: #1E3A5F; font-size: 13.5px; line-height: 1.65; }

/* ---------- 5-KPI row ---------- */
.kpi5-grid { display: flex; gap: 10px; margin-bottom: 18px; flex-wrap: wrap; }
.kpi5-card {
    flex: 1; min-width: 110px; border-radius: 12px;
    padding: 14px 10px; text-align: center; border: 2px solid;
}
.kpi5-card.green   { background: #F0FFF4; border-color: #86EFAC; }
.kpi5-card.amber   { background: #FFFBEB; border-color: #FCD34D; }
.kpi5-card.red     { background: #FFF0F0; border-color: #FCA5A5; }
.kpi5-card.neutral { background: #F0F7FF; border-color: #B3D4F0; }
.kpi5-flag  { font-size: 18px; margin-bottom: 3px; }
.kpi5-label { font-size: 10px; color: #64748B; font-weight: 600;
              text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 5px; }
.kpi5-value { font-size: 20px; font-weight: 700; color: #1E293B; line-height: 1.1; }
.kpi5-sub   { font-size: 10px; color: #64748B; margin-top: 3px; }

/* ---------- insight card ---------- */
.insight-card {
    background: linear-gradient(135deg, #F0F7FF 0%, #E8F4FE 100%);
    border: 1px solid #BFDBFE;
    border-left: 4px solid #0072BC;
    border-radius: 12px;
    padding: 18px 20px;
    margin-bottom: 8px;
}
.insight-card p { margin: 0; color: #1E3A5F; font-size: 14px; line-height: 1.6; }

/* ---------- chat styling ---------- */
.chat-header {
    background: linear-gradient(135deg, #E3F2FD, #E0F7FA);
    border: 1px solid #B3D9F0;
    border-radius: 12px;
    padding: 12px 18px;
    margin-bottom: 14px;
    display: flex;
    align-items: center;
    gap: 10px;
}
.chat-badge {
    background: linear-gradient(135deg, #0072BC, #0097A7);
    color: white;
    font-size: 11px;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
.chat-title { font-size: 14px; font-weight: 600; color: #1A237E; }
.chat-sub   { font-size: 12px; color: #546E7A; margin-top: 2px; }

/* ---------- upload panel ---------- */
.upload-panel {
    background: #F8FAFE;
    border: 1.5px dashed #90CAF9;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 14px;
}
.upload-panel-solid {
    background: #F8FAFE;
    border: 1px solid #DBEAFE;
    border-radius: 12px;
    padding: 20px;
}

/* ---------- quote analysis column backgrounds ----------
   Target Streamlit's own column vertical-block containers.
   Tab 2 (Quote Analysis) is the 2nd tab-panel child.
--------------------------------------------------------- */
[data-baseweb="tab-panel"]:nth-of-type(2)
  [data-testid="stHorizontalBlock"]
  > [data-testid="column"]:first-child
  > [data-testid="stVerticalBlock"] {
    background: #EFF6FF;
    border-radius: 16px;
    padding: 20px 18px 24px 18px !important;
    border: 1px solid #DBEAFE;
}
[data-baseweb="tab-panel"]:nth-of-type(2)
  [data-testid="stHorizontalBlock"]
  > [data-testid="column"]:last-child
  > [data-testid="stVerticalBlock"] {
    background: #F0FDF9;
    border-radius: 16px;
    padding: 20px 18px 24px 18px !important;
    border: 1px solid #CCFBF1;
}

/* ---------- line items table ---------- */
.li-table { width: 100%; border-collapse: collapse; font-size: 13px; margin-bottom: 14px; }
.li-table thead tr { background: #1E3A5F; color: #fff; }
.li-table thead th {
    padding: 9px 10px; text-align: left; font-weight: 600;
    font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;
}
.li-table thead th:nth-child(3),
.li-table thead th:nth-child(4),
.li-table thead th:nth-child(5) { text-align: right; }
.li-table tbody tr { border-bottom: 1px solid #E2E8F0; }
.li-table tbody tr:hover { filter: brightness(0.97); }
.li-table td { padding: 8px 10px; vertical-align: top; }
.li-table td:nth-child(3),
.li-table td:nth-child(4),
.li-table td:nth-child(5) { text-align: right; font-variant-numeric: tabular-nums; }
.li-row-low    { background: #F0FFF4; }
.li-row-amber  { background: #FFFBEB; }
.li-row-high   { background: #FFF1F2; }
.li-badge {
    display: inline-block; padding: 2px 8px; border-radius: 10px;
    font-size: 10px; font-weight: 700; letter-spacing: 0.3px;
}
.li-badge-low   { background: #DCFCE7; color: #166534; }
.li-badge-amber { background: #FEF3C7; color: #92400E; }
.li-badge-high  { background: #FFE4E6; color: #9F1239; }
.li-notes { font-size: 11px; color: #64748B; margin-top: 2px; }

/* ---------- status badges ---------- */
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: 600;
}
.badge-accepted { background: #DCFCE7; color: #166534; }
.badge-rejected { background: #FEE2E2; color: #991B1B; }
.badge-pending  { background: #FEF9C3; color: #854D0E; }

/* ---------- tab override ---------- */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    background: #EFF6FF;
    border-radius: 10px;
    padding: 4px;
    border: 1px solid #DBEAFE;
    margin-bottom: 20px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    padding: 8px 20px;
    font-weight: 500;
    font-size: 14px;
    color: #475569;
    border: none;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #0072BC, #0097A7) !important;
    color: white !important;
}

/* ---------- button override ---------- */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #0072BC, #0097A7);
    border: none;
    border-radius: 8px;
    font-weight: 600;
    letter-spacing: 0.3px;
    transition: opacity 0.2s;
}
.stButton > button[kind="primary"]:hover { opacity: 0.88; }

/* ---------- form submit button ---------- */
.stFormSubmitButton > button {
    border-radius: 8px;
    font-weight: 600;
}

/* ---------- expander ---------- */
.streamlit-expanderHeader {
    background: #F0F7FF;
    border-radius: 8px;
    font-weight: 500;
}

/* ---------- dataframe ---------- */
.stDataFrame { border-radius: 10px; overflow: hidden; }

/* ---------- source panel ---------- */
.src-panel-title {
    font-size: 11px;
    font-weight: 700;
    color: #94A3B8;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 14px;
    padding-bottom: 8px;
    border-bottom: 1px solid #E2ECF8;
}
.src-file-card {
    background: #fff;
    border: 1px solid #E2ECF8;
    border-radius: 10px;
    margin-bottom: 12px;
    overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
}
.src-file-header {
    background: linear-gradient(90deg, #EEF4FF 0%, #F0F7FF 100%);
    padding: 9px 14px;
    font-size: 12px;
    font-weight: 600;
    color: #0072BC;
    border-bottom: 1px solid #E2ECF8;
    display: flex;
    align-items: center;
    gap: 6px;
}
.src-file-body {
    padding: 10px 14px;
    font-size: 11px;
    color: #475569;
    line-height: 1.6;
    max-height: 160px;
    overflow-y: auto;
    white-space: pre-wrap;
    font-family: 'Inter', monospace;
}
.src-quote-card {
    background: #fff;
    border: 1px solid #E2ECF8;
    border-left: 3px solid #0097A7;
    border-radius: 10px;
    padding: 11px 14px;
    margin-bottom: 10px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
}
.src-quote-id {
    font-weight: 700;
    color: #0072BC;
    font-size: 13px;
    margin-bottom: 5px;
}
.src-quote-field {
    font-size: 12px;
    color: #64748B;
    margin-top: 3px;
    line-height: 1.4;
}
.src-empty {
    text-align: center;
    padding: 50px 20px;
    color: #CBD5E1;
}
.src-cited-badge {
    display: inline-block;
    background: #0097A7;
    color: #fff;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 0.06em;
    padding: 2px 7px;
    border-radius: 20px;
    margin-left: 8px;
    vertical-align: middle;
    text-transform: uppercase;
}

/* ---------- hide "press enter to submit" hint ---------- */
[data-testid="InputInstructions"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

client = get_client()


def _stream(api_stream):
    for chunk in api_stream:
        if chunk.choices and chunk.choices[0].delta.content:
            yield chunk.choices[0].delta.content


_DB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database")


try:
    import fitz as _fitz
except ImportError:
    _fitz = None


def _extract_highlight_terms_for_pdf(response_text: str, pdf_text: str) -> list[str]:
    """Find phrases from the AI response that appear verbatim in the PDF, plus key values."""
    terms: set[str] = set()
    pdf_lower = pdf_text.lower()

    # Sliding window: take n-grams from the response, keep those that exist in the PDF
    words = response_text.split()
    for window in range(8, 2, -1):
        for i in range(len(words) - window + 1):
            phrase = " ".join(words[i : i + window])
            if len(phrase) >= 10 and phrase.lower() in pdf_lower:
                terms.add(phrase)

    # Always highlight monetary amounts, large numbers, and dates from the response
    terms.update(re.findall(r'£[\d,]+(?:\.\d+)?', response_text))
    terms.update(re.findall(r'€[\d,]+(?:\.\d+)?', response_text))
    terms.update(re.findall(r'\b\d{1,3}(?:,\d{3})+(?:\.\d+)?\b', response_text))
    terms.update(re.findall(r'\b\d{4}-\d{2}-\d{2}\b', response_text))

    return [t for t in terms if len(t) >= 3]


def _render_pdf_pages(pdf_path: str, terms: list[str]) -> list[bytes] | None:
    """Return list of PNG bytes (one per page) with highlights, or None if fitz unavailable."""
    if _fitz is None:
        return None
    try:
        doc = _fitz.open(pdf_path)
        for page in doc:
            for term in terms:
                for rect in page.search_for(term):
                    annot = page.add_highlight_annot(rect)
                    annot.set_colors(stroke=(1.0, 0.92, 0.23))
                    annot.update()
        mat = _fitz.Matrix(1.8, 1.8)
        return [page.get_pixmap(matrix=mat, alpha=False).tobytes("png") for page in doc]
    except Exception:
        return None


_PDFJS_VERSION = "3.11.174"
_PDFJS_CDN = f"https://cdnjs.cloudflare.com/ajax/libs/pdf.js/{_PDFJS_VERSION}"


def _pdfjs_preview_html(pdf_bytes: bytes, height: int = 500) -> str:
    """Render PDF bytes in-browser via PDF.js (no highlights)."""
    b64 = base64.b64encode(pdf_bytes).decode()
    return f"""<!DOCTYPE html>
<html>
<head>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #F8FAFE; font-family: sans-serif; }}
  #viewer {{ padding: 8px; display: flex; flex-direction: column; gap: 8px; }}
  canvas {{ display: block; width: 100%; box-shadow: 0 1px 4px rgba(0,0,0,.18); background: #fff; border-radius: 3px; }}
  #msg {{ padding: 12px; color: #64748B; font-size: 13px; }}
</style>
</head>
<body>
  <div id="viewer"><div id="msg">Loading PDF…</div></div>
<script src="{_PDFJS_CDN}/pdf.min.js"></script>
<script>
pdfjsLib.GlobalWorkerOptions.workerSrc = '{_PDFJS_CDN}/pdf.worker.min.js';
const data = atob('{b64}');
const buf = new Uint8Array(data.length);
for (let i = 0; i < data.length; i++) buf[i] = data.charCodeAt(i);
pdfjsLib.getDocument({{ data: buf }}).promise.then(pdf => {{
  document.getElementById('msg').remove();
  for (let p = 1; p <= pdf.numPages; p++) {{
    pdf.getPage(p).then(page => {{
      const vp = page.getViewport({{ scale: 1.6 }});
      const canvas = document.createElement('canvas');
      canvas.height = vp.height;
      canvas.width  = vp.width;
      document.getElementById('viewer').appendChild(canvas);
      page.render({{ canvasContext: canvas.getContext('2d'), viewport: vp }});
    }});
  }}
}}).catch(e => {{ document.getElementById('msg').textContent = 'Preview error: ' + e.message; }});
</script>
</body>
</html>"""


def _pdfjs_html(pdf_path: str, terms: list[str]) -> str:
    """Return an HTML string that renders the PDF with PDF.js and highlights terms."""
    with open(pdf_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    # Build a JS array of lowercase search strings
    terms_js = json.dumps([t.lower() for t in terms if len(t) >= 3])
    return f"""<!DOCTYPE html>
<html>
<head>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #F8FAFE; font-family: sans-serif; }}
  #viewer {{ padding: 8px; display: flex; flex-direction: column; gap: 8px; }}
  canvas {{ display: block; width: 100%; box-shadow: 0 1px 4px rgba(0,0,0,.18); background: #fff; }}
  #msg {{ padding: 12px; color: #64748B; font-size: 13px; }}
</style>
</head>
<body>
  <div id="viewer"><div id="msg">Loading PDF…</div></div>
<script src="{_PDFJS_CDN}/pdf.min.js"></script>
<script>
pdfjsLib.GlobalWorkerOptions.workerSrc = '{_PDFJS_CDN}/pdf.worker.min.js';
const terms = {terms_js};

function highlight(ctx, vp, textItems) {{
  if (!terms.length) return;
  ctx.save();
  ctx.globalAlpha = 0.35;
  ctx.fillStyle = '#FFE234';
  textItems.forEach(item => {{
    const s = item.str.toLowerCase();
    terms.forEach(t => {{
      let idx = s.indexOf(t);
      while (idx !== -1) {{
        const pre = item.str.slice(0, idx);
        const charW = item.width / (item.str.length || 1);
        const x = item.transform[4] + pre.length * charW;
        const y = vp.height - item.transform[5];
        const w = t.length * charW;
        const h = item.height || 12;
        ctx.fillRect(x * vp.scale, (y - h * 0.85) * vp.scale, w * vp.scale, h * vp.scale);
        idx = s.indexOf(t, idx + 1);
      }}
    }});
  }});
  ctx.restore();
}}

const b64 = "{b64}";
const bin = atob(b64);
const bytes = new Uint8Array(bin.length);
for (let i = 0; i < bin.length; i++) bytes[i] = bin.charCodeAt(i);

pdfjsLib.getDocument({{data: bytes}}).promise.then(pdf => {{
  document.getElementById('msg').remove();
  const render = n => pdf.getPage(n).then(page => {{
    const vp = page.getViewport({{scale: 1.6}});
    const canvas = document.createElement('canvas');
    canvas.width = vp.width;
    canvas.height = vp.height;
    document.getElementById('viewer').appendChild(canvas);
    const ctx = canvas.getContext('2d');
    page.render({{canvasContext: ctx, viewport: vp}}).promise.then(() =>
      page.getTextContent().then(tc => {{
        highlight(ctx, vp, tc.items);
        if (n < pdf.numPages) render(n + 1);
      }})
    );
  }});
  render(1);
}}).catch(e => {{
  document.getElementById('viewer').innerHTML =
    '<div id="msg">Could not render PDF: ' + e.message + '</div>';
}});
</script>
</body>
</html>"""


def _build_line_items_excel(line_items: list[dict], parsed: dict, analysis: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Line Items"

    # ── Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Quote Analysis — {parsed.get('supplier', 'Unknown')} — {parsed.get('description', '')}"
    ws["A1"].font = Font(bold=True, size=12, color="1A237E")
    ws["A1"].fill = PatternFill("solid", fgColor="E3F2FD")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header row
    headers = ["#", "Description", "Unit", "Quoted (£)", "% of Total",
               "Market Est. (£)", "Variance (£)", "Variance (%)", "Risk", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0072BC")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    fills = {
        "Low":   PatternFill("solid", fgColor="DCFCE7"),
        "Amber": PatternFill("solid", fgColor="FEF9C3"),
        "High":  PatternFill("solid", fgColor="FEE2E2"),
    }
    total_quoted = sum(item.get("amount", 0) for item in line_items) or 1

    for i, item in enumerate(line_items, 1):
        row = i + 2
        amount = item.get("amount", 0)
        mkt    = item.get("market_estimate", 0)
        var    = amount - mkt
        var_pct = (var / mkt) if mkt > 0 else 0
        pct_tot = amount / total_quoted
        risk = item.get("risk_level", "Low")
        fill = fills.get(risk, fills["Low"])

        values = [i, item.get("description", ""), item.get("unit", "lump sum"),
                  amount, pct_tot, mkt, var, var_pct, risk, item.get("notes", "")]
        fmts   = [None, None, None, '£#,##0', '0.0%', '£#,##0', '£#,##0', '0.0%', None, None]
        aligns = ["center","left","center","center","center","center","center","center","center","left"]

        for col, (val, fmt, aln) in enumerate(zip(values, fmts, aligns), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal=aln, vertical="center", wrap_text=(col == 10))
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[row].height = 16

    # ── Totals row
    tr = len(line_items) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    tc = ws.cell(row=tr, column=4, value=total_quoted)
    tc.font = Font(bold=True); tc.number_format = '£#,##0'
    tp = ws.cell(row=tr, column=5, value=1.0)
    tp.font = Font(bold=True); tp.number_format = '0.0%'

    for col, w in enumerate([4, 36, 14, 14, 12, 14, 13, 12, 9, 46], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── KPI summary sheet
    ws2 = wb.create_sheet("Benchmark KPIs")
    for col, h in enumerate(["KPI", "Submitted", "Benchmark Reference", "Status"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0072BC")
        c.alignment = Alignment(horizontal="center")

    sp = analysis.get("supplier_premium_pct")
    kpis = [
        ("vs. Market Avg",      f"{analysis['new_quote_vs_mean_pct']:+.1f}%",
         "< 0% = below market avg",
         "High" if analysis['new_quote_vs_mean_pct'] > 20 else ("Amber" if analysis['new_quote_vs_mean_pct'] > 0 else "Low")),
        ("Percentile Rank",     f"{analysis['new_quote_percentile']:.0f}th",
         "< 33rd = competitive",
         "High" if analysis['new_quote_percentile'] > 75 else ("Amber" if analysis['new_quote_percentile'] > 33 else "Low")),
        ("Z-Score",             f"{analysis['new_quote_z_score']:+.2f}σ",
         "|Z| < 1.0 = normal range",
         "High" if abs(analysis['new_quote_z_score']) > 2 else ("Amber" if abs(analysis['new_quote_z_score']) > 1 else "Low")),
        ("Supplier Premium",    f"{sp:+.1f}%" if sp is not None else "N/A",
         "< 0% = below own avg",
         "High" if (sp or 0) > 25 else ("Amber" if (sp or 0) > 0 else "Low")),
        ("vs. Historical Best", f"{analysis['best_price_distance_pct']:+.1f}%",
         "< 20% = near best price",
         "High" if analysis['best_price_distance_pct'] > 40 else ("Amber" if analysis['best_price_distance_pct'] > 20 else "Low")),
    ]
    kpi_fills = {"Low": "DCFCE7", "Amber": "FEF9C3", "High": "FEE2E2"}
    for row, (label, val, ref, status) in enumerate(kpis, 2):
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=val).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=3, value=ref)
        c = ws2.cell(row=row, column=4, value=status)
        c.fill = PatternFill("solid", fgColor=kpi_fills.get(status, "FFFFFF"))
        c.alignment = Alignment(horizontal="center")
    for col, w in enumerate([26, 16, 28, 12], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _extract_sources(response_text: str) -> dict:
    sources: dict = {"pdfs": [], "quotes": [], "response_text": response_text}
    cited_pdfs = {p.lower() for p in re.findall(r"\b([\w\-]+\.pdf)\b", response_text, re.IGNORECASE)}
    response_lower = response_text.lower()
    for fname in sorted(os.listdir(_DB_DIR)):
        if not fname.lower().endswith(".pdf"):
            continue
        fpath = os.path.join(_DB_DIR, fname)
        stem = re.sub(r"_\d{8}", "", fname[:-4]).lower()
        try:
            reader = PdfReader(fpath)
            text = "\n".join(p.extract_text() or "" for p in reader.pages).strip()
        except Exception:
            continue
        if not text:
            continue
        # Match by explicit filename/stem citation OR by content overlap (5-word phrase)
        name_match = fname.lower() in cited_pdfs or stem in response_lower
        if not name_match:
            pdf_words = text.split()
            name_match = any(
                " ".join(pdf_words[i : i + 5]).lower() in response_lower
                for i in range(0, max(1, len(pdf_words) - 4), 3)
            )
        if name_match:
            sources["pdfs"].append({"filename": fname, "text": text, "path": fpath})
    cited_ids = set(re.findall(r"\b(Q\d{3,4})\b", response_text))
    if cited_ids:
        sources["quotes"] = [q for q in load_quotes() if q.get("id") in cited_ids]
    return sources


# ── Session state ──────────────────────────────────────────────────────────────
if "messages" not in st.session_state:
    st.session_state.messages = []
if "last_sources" not in st.session_state:
    st.session_state.last_sources = {"pdfs": [], "quotes": [], "response_text": ""}
if "pending_response" not in st.session_state:
    st.session_state.pending_response = False

# ── Hero banner ────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div class="hero-title">💧 AquaCapital — Quote Intelligence Platform</div>
  <div class="hero-sub">AI-powered capital project procurement analysis for UK water utilities · AMP8 benchmarked</div>
  <div class="hero-pills">
    <span class="hero-pill">🏗️ Capital Projects</span>
    <span class="hero-pill">📊 AMP8 Benchmarks</span>
    <span class="hero-pill">🤖 AI Insights</span>
    <span class="hero-pill">💷 GBP · Inflation-Adjusted</span>
    <span class="hero-pill">📄 PDF Quote Analysis</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["💬  Quote Chatbot", "📊  Quote Analysis"])


# ── Tab 1 : Chatbot ────────────────────────────────────────────────────────────
with tab1:
    chat_col, src_col = st.columns([3, 2], gap="large")

    with chat_col:
        st.markdown("""
        <div class="chat-header">
          <div>
            <div style="display:flex;align-items:center;gap:8px;">
              <span class="chat-title">Procurement Intelligence Assistant</span>
              <span class="chat-badge">AI</span>
            </div>
            <div class="chat-sub">Ask about historical quotes, contractor performance, pricing trends, AMP8 benchmarks&hellip;</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("chat_form", clear_on_submit=True):
            col_input, col_btn, col_clear = st.columns([5, 2, 1.5])
            with col_input:
                prompt = st.text_input(
                    "prompt",
                    placeholder="e.g. What did we pay for ductile iron pipes? Which contractor quoted lowest on pipeline work?",
                    label_visibility="collapsed",
                )
            with col_btn:
                submitted = st.form_submit_button("Send ➤", use_container_width=True)
            with col_clear:
                cleared = st.form_submit_button("Clear", use_container_width=True)

        if cleared:
            st.session_state.messages = []
            st.session_state.last_sources = {"pdfs": [], "quotes": [], "response_text": ""}
            st.session_state.pending_response = False
            st.rerun()

        # Stage 1: store user message immediately, then rerun to show it
        if submitted and prompt.strip():
            st.session_state.messages.append({"role": "user", "content": prompt.strip()})
            st.session_state.pending_response = True
            st.rerun()

        # Show all messages (user message visible straight away)
        for msg in reversed(st.session_state.messages):
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        if not st.session_state.messages:
            st.markdown("""
            <div style="text-align:center;padding:40px 20px;color:#94A3B8;">
              <div style="font-size:40px;margin-bottom:12px;">💬</div>
              <div style="font-size:15px;font-weight:500;color:#64748B;">No conversation yet</div>
              <div style="font-size:13px;margin-top:6px;">Try asking about quotes, contractors, or price trends</div>
            </div>
            """, unsafe_allow_html=True)

        # Stage 2: call API after user message is rendered
        if st.session_state.pending_response:
            with st.spinner("Analysing your question…"):
                last_user_msg = next(
                    (m["content"] for m in reversed(st.session_state.messages) if m["role"] == "user"),
                    "",
                )
                system_prompt = build_system_prompt(query=last_user_msg)
                api_messages = [{"role": "system", "content": system_prompt}] + [
                    {"role": m["role"], "content": m["content"]}
                    for m in st.session_state.messages
                ]
                stream = client.chat.completions.create(
                    model=DEPLOYMENT,
                    messages=api_messages,
                    stream=True,
                    temperature=0.3,
                    max_completion_tokens=800,
                )
                response = "".join(_stream(stream))
            st.session_state.messages.append({"role": "assistant", "content": response})
            st.session_state.last_sources = _extract_sources(response)
            st.session_state.pending_response = False
            st.rerun()

    with src_col:
        sources = st.session_state.last_sources
        has_sources = bool(sources["pdfs"] or sources["quotes"])

        st.markdown('<div class="src-panel-title">📎 &nbsp;Sources</div>', unsafe_allow_html=True)

        if not has_sources:
            st.markdown("""
            <div class="src-empty">
              <div style="font-size:32px;margin-bottom:10px;">📂</div>
              <div style="font-size:13px;font-weight:500;color:#94A3B8;">No sources yet</div>
              <div style="font-size:12px;margin-top:6px;color:#CBD5E1;">Files cited by the AI will<br>appear here after your first question.</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            response_text = sources.get("response_text", "")

            for pdf in sources["pdfs"]:
                st.markdown(
                    f'<div class="src-file-header">📄 &nbsp;{pdf["filename"]}</div>',
                    unsafe_allow_html=True,
                )
                terms = _extract_highlight_terms_for_pdf(response_text, pdf["text"])
                pages = _render_pdf_pages(pdf["path"], terms)
                if pages:
                    for img_bytes in pages:
                        st.image(img_bytes, use_container_width=True)
                else:
                    # Fallback: render via PDF.js in the browser (no native deps required)
                    html = _pdfjs_html(pdf["path"], terms)
                    components.html(html, height=600, scrolling=True)

            for q in sources["quotes"]:
                price = q.get("total_price", 0)
                price_fmt = f"£{price:,.0f}" if isinstance(price, (int, float)) else str(price)
                st.markdown(f"""
                <div class="src-quote-card">
                  <div class="src-quote-id">{q.get("id", "—")}</div>
                  <div class="src-quote-field">🏢 &nbsp;{q.get("supplier", "—")}</div>
                  <div class="src-quote-field">📋 &nbsp;{q.get("description", "—")}</div>
                  <div class="src-quote-field">💷 &nbsp;{price_fmt} &nbsp;·&nbsp; {q.get("date", "—")}</div>
                  <div class="src-quote-field">🏷️ &nbsp;{q.get("category", "—")}</div>
                </div>
                """, unsafe_allow_html=True)


# ── Tab 2 : Quote Analysis ─────────────────────────────────────────────────────
with tab2:
    col_upload, col_results = st.columns([1, 1.4], gap="large")

    # ── Left panel: upload ─────────────────────────────────────────────────────
    with col_upload:
        st.markdown("""
        <div class="section-header">
          <div class="section-icon blue">📥</div>
          <div>
            <div class="section-title">Submit a Quote for Analysis</div>
          </div>
        </div>
        <div class="section-divider"></div>
        """, unsafe_allow_html=True)

        quote_text = ""
        uploaded = st.file_uploader(
            "Upload a quote document",
            type=["pdf"],
            label_visibility="collapsed",
        )
        if uploaded:
            try:
                raw_bytes = uploaded.read()
                if len(raw_bytes) == 0:
                    st.error("The uploaded file is empty.")
                elif uploaded.name.lower().endswith(".pdf"):
                    reader = PdfReader(io.BytesIO(raw_bytes))
                    quote_text = "\n".join(
                        page.extract_text() or "" for page in reader.pages
                    ).strip()
                    if not quote_text:
                        st.error("Could not extract text from this PDF — it may be image-based.")
                    else:
                        st.success(f"✅ PDF ready — {len(quote_text):,} characters extracted")
                        with st.expander("📄 PDF preview", expanded=True):
                            components.html(
                                _pdfjs_preview_html(raw_bytes),
                                height=520,
                                scrolling=True,
                            )
                else:
                    quote_text = raw_bytes.decode("utf-8")
                    with st.expander("📄 File content preview", expanded=False):
                        st.text_area("", quote_text, height=180, label_visibility="collapsed")
            except Exception as e:
                st.error(f"Could not read file: {e}")

        st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)
        run_btn = st.button(
            "🔍  Run Analysis",
            type="primary",
            use_container_width=True,
            disabled=not bool(quote_text.strip()),
        )

        if run_btn and quote_text.strip():
            parsed = None
            with st.spinner("Extracting structured data from quote…"):
                try:
                    parsed = parse_quote(quote_text)
                except json.JSONDecodeError:
                    st.error("Could not parse quote into structured data. Try adding more detail.")
                except Exception as e:
                    st.error(f"Parsing error: {e}")

            if parsed:
                with st.expander("✅ Parsed quote data", expanded=False):
                    st.json(parsed)
                comparable_result = {}
                with st.spinner("Finding comparable quotes in database…"):
                    try:
                        comparable_result = find_comparable_quotes(quote_text, parsed)
                    except Exception:
                        comparable_result = {}
                comparable_ids = comparable_result.get("comparable_ids") or None
                with st.spinner("Benchmarking against comparable quotes…"):
                    analysis = analyze_quote(parsed, comparable_ids=comparable_ids)
                analysis["comparable_reasoning"] = comparable_result.get("reasoning", "")
                line_items = []
                if "error" not in analysis:
                    with st.spinner("Extracting line items…"):
                        try:
                            line_items = parse_line_items(
                                quote_text,
                                parsed.get("description", ""),
                                analysis["new_quote_price"],
                                analysis["inflation_adjusted_mean"],
                            )
                        except Exception:
                            line_items = []
                    with st.spinner("Running procurement agent analysis…"):
                        try:
                            agent_data = build_agent_summary(parsed, analysis, line_items)
                        except Exception:
                            agent_data = {}
                else:
                    agent_data = {}
                st.session_state["analysis"] = analysis
                st.session_state["parsed_quote"] = parsed
                st.session_state["line_items"] = line_items
                st.session_state["agent_summary"] = agent_data
                st.rerun()

        # ── Database overview when idle ────────────────────────────────────────
        if "analysis" not in st.session_state:
            st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="section-header">
              <div class="section-icon teal">🗄️</div>
              <div><div class="section-title">Database Snapshot</div></div>
            </div>
            <div class="section-divider"></div>
            """, unsafe_allow_html=True)
            try:
                df = get_dataframe()
                total_spend  = df["total_price"].sum()
                n_quotes     = len(df)
                n_suppliers  = df["supplier"].nunique()
                accepted_pct = (df["status"] == "accepted").mean() * 100 if "status" in df.columns else 0

                st.markdown(f"""
                <div class="kpi5-grid">
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">💷</div>
                    <div class="kpi5-label">Total Spend</div>
                    <div class="kpi5-value">£{total_spend/1e6:.1f}M</div>
                    <div class="kpi5-sub">across all quotes</div>
                  </div>
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">📋</div>
                    <div class="kpi5-label">Quotes</div>
                    <div class="kpi5-value">{n_quotes}</div>
                    <div class="kpi5-sub">in database</div>
                  </div>
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">🏢</div>
                    <div class="kpi5-label">Contractors</div>
                    <div class="kpi5-value">{n_suppliers}</div>
                    <div class="kpi5-sub">unique suppliers</div>
                  </div>
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">✅</div>
                    <div class="kpi5-label">Accepted</div>
                    <div class="kpi5-value">{accepted_pct:.0f}%</div>
                    <div class="kpi5-sub">acceptance rate</div>
                  </div>
                </div>
                """, unsafe_allow_html=True)
            except Exception:
                pass

    # ── Right panel: results ───────────────────────────────────────────────────
    with col_results:
        if "analysis" in st.session_state:
            analysis   = st.session_state["analysis"]
            parsed     = st.session_state["parsed_quote"]
            agent_data = st.session_state.get("agent_summary", {})
            line_items = st.session_state.get("line_items", [])

            if "error" in analysis:
                st.error(analysis["error"])
            else:
                # ── Agent Summary (top of dashboard) ──────────────────────────
                summary_md = agent_data.get("summary", "")
                if summary_md:
                    st.markdown("""
                    <div class="agent-summary-card">
                      <div class="agent-summary-header">
                        <span class="agent-summary-title">🤖 Procurement Agent Summary</span>
                        <span class="agent-summary-badge">AI ANALYSIS</span>
                      </div>
                    </div>""", unsafe_allow_html=True)
                    st.markdown(summary_md)

                # ── 5 KPI cards ────────────────────────────────────────────────
                st.markdown("""
                <div class="section-header">
                  <div class="section-icon blue">📊</div>
                  <div><div class="section-title">Benchmark KPIs</div></div>
                </div>
                <div class="section-divider"></div>
                """, unsafe_allow_html=True)

                def _kpi_class(val, thresholds, invert=False):
                    lo, hi = thresholds
                    if invert:
                        return "green" if val <= lo else ("amber" if val <= hi else "red")
                    return "green" if val < lo else ("amber" if val < hi else "red")

                pct        = analysis["new_quote_vs_mean_pct"]
                rank       = analysis["new_quote_percentile"]
                z          = analysis["new_quote_z_score"]
                sp         = analysis.get("supplier_premium_pct")
                best_dist  = analysis["best_price_distance_pct"]

                c1 = _kpi_class(pct,       (0, 20))
                c2 = _kpi_class(rank,      (33, 75))
                c3 = _kpi_class(abs(z),    (1, 2))
                c4 = _kpi_class(sp or 0,   (0, 25)) if sp is not None else "neutral"
                c5 = _kpi_class(best_dist, (20, 40))

                flag = {"green": "✅", "amber": "⚠️", "red": "🔴", "neutral": "➖"}
                sp_val = f"{sp:+.1f}%" if sp is not None else "N/A"
                sp_sub = "vs supplier's own avg" if sp is not None else "no prior data"

                st.markdown(f"""
                <div class="kpi5-grid">
                  <div class="kpi5-card {c1}">
                    <div class="kpi5-flag">{flag[c1]}</div>
                    <div class="kpi5-label">vs Market Avg</div>
                    <div class="kpi5-value">{pct:+.1f}%</div>
                    <div class="kpi5-sub">inflation-adj. avg</div>
                  </div>
                  <div class="kpi5-card {c2}">
                    <div class="kpi5-flag">{flag[c2]}</div>
                    <div class="kpi5-label">Percentile Rank</div>
                    <div class="kpi5-value">{rank:.0f}<sup style="font-size:12px">th</sup></div>
                    <div class="kpi5-sub">of {analysis['sample_size']} quotes</div>
                  </div>
                  <div class="kpi5-card {c3}">
                    <div class="kpi5-flag">{flag[c3]}</div>
                    <div class="kpi5-label">Z-Score</div>
                    <div class="kpi5-value">{z:+.2f}<span style="font-size:13px">σ</span></div>
                    <div class="kpi5-sub">statistical deviation</div>
                  </div>
                  <div class="kpi5-card {c4}">
                    <div class="kpi5-flag">{flag[c4]}</div>
                    <div class="kpi5-label">Supplier Premium</div>
                    <div class="kpi5-value">{sp_val}</div>
                    <div class="kpi5-sub">{sp_sub}</div>
                  </div>
                  <div class="kpi5-card {c5}">
                    <div class="kpi5-flag">{flag[c5]}</div>
                    <div class="kpi5-label">vs Best Price</div>
                    <div class="kpi5-value">{best_dist:+.1f}%</div>
                    <div class="kpi5-sub">above historical best</div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # ── Additional KPIs from agent ─────────────────────────────────
                extra_kpis = agent_data.get("additional_kpis", [])
                if extra_kpis:
                    cards_html = "".join(f"""
                      <div class="kpi5-card {k.get('status','neutral')}">
                        <div class="kpi5-flag">{flag.get(k.get('status','neutral'),'➖')}</div>
                        <div class="kpi5-label">{k.get('label','')}</div>
                        <div class="kpi5-value" style="font-size:17px">{k.get('value','')}</div>
                        <div class="kpi5-sub">{k.get('note','')}</div>
                      </div>""" for k in extra_kpis)
                    st.markdown(f'<div class="kpi5-grid">{cards_html}</div>',
                                unsafe_allow_html=True)

                # ── Contractor benchmark chart ─────────────────────────────────
                st.markdown("""
                <div class="section-header">
                  <div class="section-icon teal">🏗️</div>
                  <div><div class="section-title">Contractor Benchmark (Inflation-Adjusted)</div></div>
                </div>
                <div class="section-divider"></div>
                """, unsafe_allow_html=True)

                bench_df = pd.DataFrame(analysis["benchmark"])
                colors_bar = [
                    "#0097A7" if s != parsed.get("supplier", "") else "#FF6B35"
                    for s in bench_df["supplier"]
                ]
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=bench_df["supplier"],
                    y=bench_df["avg_adjusted"],
                    marker_color=colors_bar,
                    text=bench_df["avg_adjusted"].apply(lambda v: f"£{v:,.0f}"),
                    textposition="outside",
                    textfont=dict(size=11),
                ))
                fig.add_hline(
                    y=analysis["new_quote_price"],
                    line_dash="dash", line_color="#E53935", line_width=2,
                    annotation_text=f"  Submitted  £{analysis['new_quote_price']:,.0f}",
                    annotation_position="top left",
                    annotation_font=dict(color="#E53935", size=11),
                )
                fig.update_layout(
                    plot_bgcolor="#F8FAFE", paper_bgcolor="#F8FAFE",
                    xaxis=dict(tickangle=-20, tickfont=dict(size=11), gridcolor="#E2ECF8"),
                    yaxis=dict(title="Price (GBP)", tickfont=dict(size=11),
                               gridcolor="#E2ECF8", tickprefix="£"),
                    height=290, margin=dict(t=30, b=10, l=10, r=10), showlegend=False,
                )
                st.plotly_chart(fig, use_container_width=True)

                # ── Historical price distribution ──────────────────────────────
                st.markdown("""
                <div class="section-header">
                  <div class="section-icon green">📈</div>
                  <div><div class="section-title">Historical Price Distribution</div></div>
                </div>
                <div class="section-divider"></div>
                """, unsafe_allow_html=True)

                hist_df = pd.DataFrame(analysis["historical_quotes"])
                fig2 = go.Figure()
                fig2.add_trace(go.Histogram(
                    x=hist_df["adjusted_price"], nbinsx=12,
                    marker_color="#0097A7", opacity=0.75, name="Historical (adj.)",
                ))
                fig2.add_vline(x=analysis["new_quote_price"],
                    line_dash="dash", line_color="#E53935", line_width=2,
                    annotation_text="Submitted", annotation_position="top right",
                    annotation_font=dict(color="#E53935", size=11))
                fig2.add_vline(x=analysis["inflation_adjusted_mean"],
                    line_dash="dot", line_color="#0072BC", line_width=2,
                    annotation_text="Avg", annotation_position="top left",
                    annotation_font=dict(color="#0072BC", size=11))
                fig2.update_layout(
                    plot_bgcolor="#F8FAFE", paper_bgcolor="#F8FAFE",
                    xaxis=dict(title="Inflation-Adjusted Price (GBP)", tickprefix="£",
                               tickfont=dict(size=11), gridcolor="#E2ECF8"),
                    yaxis=dict(title="Count", tickfont=dict(size=11), gridcolor="#E2ECF8"),
                    height=220, margin=dict(t=20, b=10, l=10, r=10), showlegend=False,
                )
                st.plotly_chart(fig2, use_container_width=True)

                # ── Outlier alert ──────────────────────────────────────────────
                if analysis["outliers"]:
                    st.warning(f"⚠️  {len(analysis['outliers'])} outlier(s) detected in historical data.")
                    with st.expander("View outlier records"):
                        st.dataframe(pd.DataFrame(analysis["outliers"]), use_container_width=True)
                else:
                    st.success("✅  No outliers detected in historical benchmark data.")

                # ── Historical quotes table ────────────────────────────────────
                reasoning = analysis.get("comparable_reasoning", "")
                ids_label = ", ".join(analysis.get("comparable_ids", []))
                if reasoning:
                    st.info(f"🔍 **Comparable quotes used:** {ids_label}  \n{reasoning}")
                with st.expander(
                    f"📋  Comparable quotes  ({analysis['sample_size']} record(s))"
                ):
                    st.dataframe(pd.DataFrame(analysis["historical_quotes"]), use_container_width=True)

                # ── Line items Excel download ──────────────────────────────────
                st.markdown("""
                <div class="section-header" style="margin-top:8px">
                  <div class="section-icon green">📥</div>
                  <div><div class="section-title">Line Items Analysis</div></div>
                </div>
                <div class="section-divider"></div>
                """, unsafe_allow_html=True)

                if line_items:
                    _risk_row  = {"Low": "li-row-low", "Amber": "li-row-amber", "High": "li-row-high"}
                    _risk_badge = {"Low": "li-badge-low", "Amber": "li-badge-amber", "High": "li-badge-high"}
                    _risk_icon  = {"Low": "✅", "Amber": "⚠️", "High": "🔴"}
                    rows_html = ""
                    for it in line_items:
                        risk   = it.get("risk_level", "Low")
                        amt    = it.get("amount", 0) or 0
                        mkt    = it.get("market_estimate", 0) or 0
                        var    = amt - mkt
                        var_sign = "+" if var >= 0 else ""
                        row_cls  = _risk_row.get(risk, "li-row-low")
                        bdg_cls  = _risk_badge.get(risk, "li-badge-low")
                        icon     = _risk_icon.get(risk, "")
                        rows_html += f"""
                        <tr class="{row_cls}">
                          <td>
                            <div style="font-weight:500;color:#1E293B">{it.get('description','')}</div>
                            <div class="li-notes">{it.get('notes','')}</div>
                          </td>
                          <td style="color:#64748B;font-size:12px">{it.get('unit','')}</td>
                          <td>£{amt:,.0f}</td>
                          <td>£{mkt:,.0f}</td>
                          <td style="{'color:#DC2626;font-weight:600' if var > 0 else 'color:#16A34A;font-weight:600'}">{var_sign}£{var:,.0f}</td>
                          <td><span class="li-badge {bdg_cls}">{icon} {risk}</span></td>
                        </tr>"""
                    st.markdown(f"""
                    <table class="li-table">
                      <thead>
                        <tr>
                          <th style="width:34%">Description</th>
                          <th style="width:10%">Unit</th>
                          <th style="width:13%">Quoted (£)</th>
                          <th style="width:14%">Market Est. (£)</th>
                          <th style="width:13%">Variance (£)</th>
                          <th style="width:10%">Risk</th>
                        </tr>
                      </thead>
                      <tbody>{rows_html}</tbody>
                    </table>
                    """, unsafe_allow_html=True)

                    excel_bytes = _build_line_items_excel(line_items, parsed, analysis)
                    supplier_slug = re.sub(r"[^\w]", "_", parsed.get("supplier", "quote"))
                    st.download_button(
                        "📥  Download Line Items Analysis (Excel)",
                        data=excel_bytes,
                        file_name=f"quote_analysis_{supplier_slug}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    st.caption("Line item breakdown unavailable for this quote.")

                st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
                if st.button("🔄  Clear analysis", use_container_width=False):
                    for k in ("analysis", "parsed_quote", "line_items", "agent_summary"):
                        st.session_state.pop(k, None)
                    st.rerun()

        else:
            # ── Overview charts ────────────────────────────────────────────────
            st.markdown("""
            <div class="section-header">
              <div class="section-icon teal">📊</div>
              <div><div class="section-title">Portfolio Overview</div></div>
            </div>
            <div class="section-divider"></div>
            """, unsafe_allow_html=True)

            try:
                df = get_dataframe()
                total_spend = df["total_price"].sum()
                n_quotes    = len(df)
                n_suppliers = df["supplier"].nunique()
                accepted_pct = (df["status"] == "accepted").mean() * 100

                st.markdown(f"""
                <div class="kpi5-grid">
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">💷</div>
                    <div class="kpi5-label">Total Portfolio Spend</div>
                    <div class="kpi5-value">£{total_spend/1e6:.1f}M</div>
                    <div class="kpi5-sub">all quotes · all years</div>
                  </div>
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">📋</div>
                    <div class="kpi5-label">Quotes on Record</div>
                    <div class="kpi5-value">{n_quotes}</div>
                    <div class="kpi5-sub">in database</div>
                  </div>
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">🏢</div>
                    <div class="kpi5-label">Contractors</div>
                    <div class="kpi5-value">{n_suppliers}</div>
                    <div class="kpi5-sub">active supply chain</div>
                  </div>
                  <div class="kpi5-card neutral">
                    <div class="kpi5-flag">✅</div>
                    <div class="kpi5-label">Acceptance Rate</div>
                    <div class="kpi5-value">{accepted_pct:.0f}%</div>
                    <div class="kpi5-sub">of all submitted quotes</div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                c1, c2 = st.columns(2)
                with c1:
                    sup_spend = (
                        df.groupby("supplier")["total_price"].sum()
                        .reset_index()
                        .rename(columns={"total_price": "Total Spend", "supplier": "Supplier"})
                        .sort_values("Total Spend", ascending=False)
                    )
                    fig = px.bar(
                        sup_spend, x="Supplier", y="Total Spend",
                        title="Total Spend by Contractor",
                        color="Total Spend",
                        color_continuous_scale=["#B3D9F0", "#0072BC", "#003B6F"],
                        text="Total Spend",
                    )
                    fig.update_traces(texttemplate="£%{text:,.0f}", textposition="outside",
                                      textfont=dict(size=9))
                    fig.update_layout(
                        plot_bgcolor="#F8FAFE", paper_bgcolor="rgba(0,0,0,0)",
                        xaxis=dict(tickangle=-20, tickfont=dict(size=9), title=""),
                        yaxis=dict(tickprefix="£", tickfont=dict(size=10),
                                   gridcolor="#E2ECF8", title=""),
                        margin=dict(t=40, b=10, l=10, r=10),
                        coloraxis_showscale=False,
                        title_font=dict(size=13, color="#1A237E"),
                    )
                    st.plotly_chart(fig, use_container_width=True)

                with c2:
                    sup_avg = (
                        df.groupby("supplier")["total_price"]
                        .mean()
                        .reset_index()
                        .rename(columns={"total_price": "Avg Quote (GBP)", "supplier": "Supplier"})
                        .sort_values("Avg Quote (GBP)", ascending=True)
                    )
                    fig2 = px.bar(
                        sup_avg,
                        x="Avg Quote (GBP)",
                        y="Supplier",
                        orientation="h",
                        title="Avg Quote Value by Contractor",
                        color="Avg Quote (GBP)",
                        color_continuous_scale=["#B3D9F0", "#0072BC", "#003B6F"],
                        text="Avg Quote (GBP)",
                    )
                    fig2.update_traces(
                        texttemplate="£%{text:,.0f}",
                        textposition="outside",
                        textfont=dict(size=10),
                    )
                    fig2.update_layout(
                        plot_bgcolor="#F8FAFE",
                        paper_bgcolor="rgba(0,0,0,0)",
                        xaxis=dict(tickprefix="£", tickfont=dict(size=10),
                                   gridcolor="#E2ECF8", title=""),
                        yaxis=dict(tickfont=dict(size=10), title=""),
                        margin=dict(t=40, b=10, l=10, r=80),
                        coloraxis_showscale=False,
                        title_font=dict(size=13, color="#1A237E"),
                    )
                    st.plotly_chart(fig2, use_container_width=True)

                # ── Spend over time ────────────────────────────────────────────
                df["year"] = df["date"].dt.year
                yearly = df.groupby("year")["total_price"].sum().reset_index()
                yearly.columns = ["Year", "Total Spend"]
                fig3 = px.area(
                    yearly, x="Year", y="Total Spend",
                    title="Annual Quote Spend",
                    color_discrete_sequence=["#0097A7"],
                )
                fig3.update_traces(fill="tozeroy", line_width=2,
                                   fillcolor="rgba(0,151,167,0.15)")
                fig3.update_layout(
                    plot_bgcolor="#F8FAFE",
                    paper_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(tickfont=dict(size=11), gridcolor="#E2ECF8", title=""),
                    yaxis=dict(tickprefix="£", tickfont=dict(size=11),
                               gridcolor="#E2ECF8", title=""),
                    height=220,
                    margin=dict(t=40, b=10, l=10, r=10),
                    title_font=dict(size=13, color="#1A237E"),
                )
                st.plotly_chart(fig3, use_container_width=True)

                # ── Quote table ────────────────────────────────────────────────
                with st.expander("📋  Full quote register", expanded=False):
                    st.dataframe(
                        df[["id", "date", "supplier", "description",
                            "total_price", "currency", "status"]]
                        .assign(date=df["date"].dt.strftime("%Y-%m-%d"))
                        .sort_values("date", ascending=False)
                        .reset_index(drop=True),
                        use_container_width=True,
                    )
            except Exception as e:
                st.error(f"Could not load database: {e}")



if __name__ == "__main__":
    import os
    import subprocess
    import sys
    import threading
    import time

    if not os.environ.get("_AQUACAPITAL_LAUNCHED"):
        import socket

        def _open_edge():
            for _ in range(30):
                time.sleep(1)
                try:
                    with socket.create_connection(("localhost", 8501), timeout=1):
                        break
                except OSError:
                    continue
            subprocess.Popen(["cmd", "/c", "start", "msedge", "http://localhost:8501"])

        threading.Thread(target=_open_edge, daemon=True).start()
        env = os.environ.copy()
        env["_AQUACAPITAL_LAUNCHED"] = "1"
        subprocess.run(
            [sys.executable, "-m", "streamlit", "run", __file__, "--server.headless=true"],
            env=env,
        )
